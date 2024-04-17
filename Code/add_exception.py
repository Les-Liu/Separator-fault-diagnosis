import pandas as pd
import math
import numpy as np
import matplotlib.pyplot as plt
import openpyxl
from openpyxl.styles import Alignment

plt.rcParams['font.sans-serif']=['SimHei']
plt.rcParams['axes.unicode_minus']=False

def ReadData(filePath):
    """
        1. 读取数据
    """
    dataSet = pd.read_excel(filePath)
    return dataSet

def AddException(data,snr):
    """
        2. 添加噪声
    """
    data = np.array(data)
    snr = 10 ** (snr / 10.0)
    xpower = np.sum(data ** 2) / len(data)
    npower = xpower / snr
    noise = np.random.randn(len(data)) * np.sqrt(npower)
    return noise+data

def EliminateNegative(data,noise_data):
    """
        3. 剔除生成噪声数据集中的负数
    """
    data = np.array(data)
    for i in range(len(noise_data)):
        if noise_data[i] < 0:
            noise_data[i] = data[i]
        else:
            continue
    return noise_data

def DrawPicture(data,noise_data,titleName):
    """
        4. 测试图
    """
    plt.figure(figsize=(20, 8), dpi=100)
    plt.plot(range(len(noise_data)), noise_data, lw=1, color="blue", label='噪声值')
    # plt.plot(range(len(data)), data, lw=4, color="yellow", label='理论值')
    plt.legend(loc='upper left', frameon=True)
    plt.title(titleName)
    plt.grid(linestyle='-.')
    plt.show()

def TestMethod(titleName,snr):
    """
        5. 调整每一个特征的参数
    """
    data = dataSet[titleName]
    noise_data = AddException(data,snr)
    noise_data = EliminateNegative(data,noise_data)
    # DrawPicture(data,noise_data,titleName=titleName)
    return noise_data

if __name__ == '__main__':
    filePath = "F:/博士资料/本地论文/自己写的论文/第三篇论文/论文/数据/Train Data/Original/WVL.xlsx"

    dataSet = ReadData(filePath)
    columnsName = dataSet.columns
    Y = dataSet["标签"]

    titleName = dataSet.columns[0]
    print("---------------------------------------",titleName,"---------------------------------------")
    noise_data_0 = TestMethod(titleName,snr=50)

    titleName = dataSet.columns[1]
    print("---------------------------------------",titleName,"---------------------------------------")
    noise_data_1 = TestMethod(titleName,snr=50)

    titleName = dataSet.columns[2]
    print("---------------------------------------",titleName,"---------------------------------------")
    noise_data_2 = TestMethod(titleName,snr=50)

    titleName = dataSet.columns[3]
    print("---------------------------------------",titleName,"---------------------------------------")
    noise_data_3 = TestMethod(titleName,snr=50)

    titleName = dataSet.columns[4]
    print("---------------------------------------",titleName,"---------------------------------------")
    noise_data_4 = TestMethod(titleName,snr=50)

    titleName = dataSet.columns[5]
    print("---------------------------------------",titleName,"---------------------------------------")
    noise_data_5 = TestMethod(titleName,snr=50)

    titleName = dataSet.columns[6]
    print("---------------------------------------",titleName,"---------------------------------------")
    noise_data_6 = TestMethod(titleName,snr=50)

    titleName = dataSet.columns[7]
    print("---------------------------------------",titleName,"---------------------------------------")
    noise_data_7 = TestMethod(titleName,snr=50)

    titleName = dataSet.columns[8]
    print("---------------------------------------",titleName,"---------------------------------------")
    noise_data_8 = TestMethod(titleName,snr=50)

    titleName = dataSet.columns[9]
    print("---------------------------------------",titleName,"---------------------------------------")
    noise_data_9 = TestMethod(titleName,snr=50)

    titleName = dataSet.columns[10]
    print("---------------------------------------",titleName,"---------------------------------------")
    noise_data_10 = TestMethod(titleName,snr=50)

    noise_dataDF = pd.DataFrame(data=[noise_data_0,noise_data_1,noise_data_2,
                                      noise_data_3,noise_data_4,noise_data_5,
                                      noise_data_6,noise_data_7,noise_data_8,
                                      noise_data_9,noise_data_10,Y]).transpose()
    noise_dataDF.columns = columnsName
    # 输出Excel
    excelFilePath = "F:/博士资料/本地论文/自己写的论文/第三篇论文/论文/数据/Train Data/Add Noise 50snr/WVL.xlsx"
    sheetName = "WVL"
    workBook = openpyxl.Workbook()
    sheet = workBook.create_sheet(title=sheetName, index=0)
    for i in range(noise_dataDF.shape[0] + 1):
        for j in range(noise_dataDF.shape[1]):
            if i == 0:
                sheet.cell(i + 1, j + 1, noise_dataDF.columns[j]).alignment = Alignment(horizontal='center',vertical='center')
            else:
                sheet.cell(i + 1, j + 1, noise_dataDF.iloc[i - 1, j]).alignment = Alignment(horizontal='center',vertical='center')
    workBook.save(excelFilePath)